from tkinter import *
from prettytable import PrettyTable
from twilio.rest import Client
from datetime import datetime
import pandas as pd
import xlsxwriter
from openpyxl import load_workbook

root=Tk()
root.title("SMS")
root.geometry("1350x750")

Name=StringVar()
Id=StringVar()
Date=StringVar()
Profit=StringVar()
Return=StringVar()
NetProfit=StringVar()
IniCap=StringVar()
NetCap=StringVar()
Number=StringVar()


def Table():
        t = PrettyTable()
        t.add_column(Name.get(),['DATE','PROFIT/LOSS',"TODAY'S RETURN",
                          'NET PROFIT(%)','INITIAL CAPITAL','NET CAPITAL'],align='l',valign='t')
        t.add_column(Id.get(),[Date.get(), Profit.get(),Return.get(),NetProfit.get(),IniCap.get(),NetCap.get()],align='r',valign='t')

        return t


def Sms():
        account_sid = ''
        auth_token = ''
        client = Client(account_sid, auth_token)
        
        message=client.messages.create(
                     body=Table(),
                     from_='',
                     to =  Number.get()
                 )

        #print(message.sid)

        print(message.status)



def append_df_to_excel(filename, df, sheet_name, startrow=None,truncate_sheet=False, **to_excel_kwargs):
    
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError
    try:     
        writer.book = load_workbook(filename)
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass
    if startrow is None:
        startrow = 0
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    writer.save()


def Save():
        df = pd.DataFrame({
                "Client's name": [Name.get()],
                "Client's ID": [Id.get()],
                "Date": [Date.get()],
                "Profit/Loss": [Profit.get()] ,
                "Today's Return On Capital": [Return.get()],
                "Net Profit": [NetProfit.get()],
                "Initial Capital": [IniCap.get()],
                "Net Capital": [NetCap.get()], })

        file=pd.ExcelFile("sms_excel.xlsx")
        worksheet=file.sheet_names
        for i in worksheet:
            if(Name.get()==i):
                append_df_to_excel('sms_excel.xlsx',df,Name.get(),index=False,header=False)
            else:
                writer = pd.ExcelWriter('sms_excel.xlsx', engine='openpyxl')

                writer.book = load_workbook('sms_excel.xlsx')
        
                writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                df.to_excel(writer, sheet_name=Name.get(),index=False)
                writer.save()
        
def Clear():
        txtName.delete(0,END)
        txtId.delete(0,END)
        txtProfit.delete(0,END)
        txtIniCap.delete(0,END)
        txtNetCap.delete(0,END)
        txtReturn.delete(0,END)
        txtReturn.insert(0,"%")
        txtNetProfit.delete(0,END)
        txtNetProfit.insert(0,"%")
        txtNumber.delete(0,END)
        txtNumber.insert(0,"+91")



#====================================Frames================================================
MainFrame= Frame(root)
MainFrame.grid()

TitleFrame = Frame(MainFrame,padx=50,pady=8)        
TitleFrame.pack(side=TOP)

lblTitle = Label(TitleFrame, bd=2,font=('arial',45,'bold'), text="SMS Information")
lblTitle.grid(sticky=W)

ButtonFrame = Frame(MainFrame,width=1350,height=70,padx=110,pady=1)        
ButtonFrame.pack(side=BOTTOM)

DataFrame = Frame(MainFrame,bd=1, width=1350,height=700,padx=100,pady=50)        
DataFrame.pack(side=BOTTOM)

#=======================Lables and entry==============================================
Label(DataFrame,text="Client's Name",padx=10,font=('arial',20,'bold')).grid(row=0,column=0,sticky=W)
txtName = Entry(DataFrame, font=('arial',20,'bold'), textvariable=Name,width=39,bd=2)
txtName.grid(row=0,column=1)

Label(DataFrame,text="Client's Id",padx=10,font=('arial',20,'bold')).grid(row=1,column=0,sticky=W)
txtId = Entry(DataFrame, font=('arial',20,'bold'), textvariable=Id,width=39,bd=2)
txtId.grid(row=1,column=1)

Label(DataFrame,text="Date",padx=10,font=('arial',20,'bold')).grid(row=2,column=0,sticky=W)
txtDate = Entry(DataFrame, font=('arial',20,'bold'), textvariable=Date,width=39,bd=2)
txtDate.grid(row=2,column=1)
now = datetime.now()
date_object = now.strftime("%d-%B-%Y")
txtDate.insert(0,date_object)

Label(DataFrame,text="Profit(+)/Loss(-)",padx=10,font=('arial',20,'bold')).grid(row=3,column=0,sticky=W)
txtProfit = Entry(DataFrame, font=('arial',20,'bold'), textvariable=Profit,width=39,bd=2)
txtProfit.grid(row=3,column=1)

Label(DataFrame,text="Toaday's Return On Capital(%)",padx=10,font=('arial',20,'bold')).grid(row=4,column=0,sticky=W)
txtReturn = Entry(DataFrame, font=('arial',20,'bold'), textvariable=Return,width=39,bd=2)
txtReturn.grid(row=4,column=1)
txtReturn.insert(0,"%")

Label(DataFrame,text="Net Profit(%)",padx=10,font=('arial',20,'bold')).grid(row=5,column=0,sticky=W)
txtNetProfit = Entry(DataFrame, font=('arial',20,'bold'), textvariable=NetProfit,width=39,bd=2)
txtNetProfit.grid(row=5,column=1)
txtNetProfit.insert(0,"%")

Label(DataFrame,text="Initial Capital",padx=10,font=('arial',20,'bold')).grid(row=6,column=0,sticky=W)
txtIniCap = Entry(DataFrame, font=('arial',20,'bold'), textvariable=IniCap,width=39,bd=2)
txtIniCap.grid(row=6,column=1)

Label(DataFrame,text="Net Capital",padx=10,font=('arial',20,'bold')).grid(row=7,column=0,sticky=W)
txtNetCap = Entry(DataFrame, font=('arial',20,'bold'), textvariable=NetCap,width=39,bd=2)
txtNetCap.grid(row=7,column=1)

Label(DataFrame,text="Client's Number",padx=10,font=('arial',20,'bold')).grid(row=8,column=0,sticky=W)
txtNumber = Entry(DataFrame, font=('arial',20,'bold'), textvariable=Number,width=39,bd=2)
txtNumber.grid(row=8,column=1)
txtNumber.insert(0,"+91")

#===================================Buttons==============================================
btnSave = Button(ButtonFrame, text="Save",font=('arial',18,'bold'),height=1,width=10,bd=5,
                activeforeground="white",activebackground="blue",command=Save)
btnSave.grid(row=0,column=0)

btnSend = Button(ButtonFrame, text="Send",font=('arial',18,'bold'),height=1,width=10,bd=5,
                activeforeground="white",activebackground="blue",command=Sms)
btnSend.grid(row=0,column=1)

btnClear = Button(ButtonFrame, text="Clear",font=('arial',18,'bold'),height=1,width=10,bd=5,
                activeforeground="white",activebackground="blue",command=Clear)
btnClear.grid(row=0,column=2)

root.mainloop()
